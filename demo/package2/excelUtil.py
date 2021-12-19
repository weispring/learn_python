import pymysql
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl



def query_all(sql):

    # 连接数据库，查询数据

    db = pymysql.connect(host='localhost',user='root',password='123456',port=3306,db='learn')

    # 使用cursor()方法获取操作游标

    cur = db.cursor()



    # 使用execute方法执行SQL语句

    cur.execute(sql)  # 返回受影响的行数

    fields = [field[0] for field in cur.description]  # 获取所有字段名

    all_data = cur.fetchall()  # 所有数据

    print(len(all_data))

    return all_data,fields



def read_mysql_to_xlsx(sql,excelname):
    # 循环数据写入内容
    jb_date_lists = query_all(sql)
    jb_date_list = jb_date_lists[0]
    descripte = jb_date_lists[1]


    #要创建的xlsx名称
    dest_filename = excelname + '.xlsx'
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "data"

    # 列名
    for i in range(0,len(descripte)):
        ws1.cell(row=1,column=i+1,value=descripte[i])


    # 写入数据
    for i in range(2,len(jb_date_list)+2):
        for j in range(0,len(descripte)):
            if jb_date_list[i-2][j] is None:
                ws1.cell(row=i, column=j+1, value='')
            else:
                ws1.cell(row=i, column=j+1, value=jb_date_list[i-2][j])

    # 创建xlsx

    wb.save(filename=dest_filename)



if __name__ == '__main__':

    sql = 'select name as "城市名称", cityID, code, name_En as "英文名称"   from t_er_city tec  where countryName   = "中国" '

    read_mysql_to_xlsx(sql,'user0205')

